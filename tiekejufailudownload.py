from __future__ import annotations

import csv
import io
import re
import time
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo

import requests

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


TIMEZONE = ZoneInfo("Europe/Vilnius")
OUTPUT_DIR = Path("downloads")
CSV_DIR = Path("csv")

SUPPLIERS = {
    "Marini": {
        "type": "xml",
        "url": "https://marini.pl/b2b/marini-b2b.xml",
        "raw_extension": ".xml",
    },
    "Zuja": {
        "type": "xml",
        "url": (
            "https://zuja.lt/index.php?route=feed/store/generate&filters=YToyOntzOjI0OiJmaWx0ZXJfY3VzdG9tZXJfZ3JvdXBfaWQiO3M6MjoiMTIiO3M6Mzoia2V5IjtzOjMyOiJjODFlNzI4ZDlkNGMyZjYzNmYwNjdmODljYzE0ODYyYyI7fQ==&key=c81e728d9d4c2f636f067f89cc14862c"
        ),
        "raw_extension": ".xml",
    },
    "Tylla": {
        "type": "xlsx",
        "url": "https://docs.google.com/spreadsheets/d/1T4Gpfk4Uv9FuvDO0WQeRfPIvqjfMxOdGIXLoKB8kE6o/export?format=xlsx",
        "raw_extension": ".xlsx",
    },
}


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]+', "_", name).strip()


def current_timestamp() -> str:
    return datetime.now(TIMEZONE).strftime("%Y-%m-%d_%H-%M-%S")


def build_filename(supplier_name: str, extension: str, timestamp: str | None = None) -> str:
    ts = timestamp or current_timestamp()
    return f"{sanitize_filename(supplier_name)}_{ts}{extension}"


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def looks_like_xml(content: bytes) -> bool:
    head = content[:500].lstrip()
    return head.startswith(b"<?xml") or head.startswith(b"<")


def download_with_retries(
    supplier_name: str,
    url: str,
    target_path: Path,
    expect_xml: bool = False,
    max_attempts: int = 4,
) -> bytes:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "*/*",
        "Connection": "keep-alive",
    }

    session = requests.Session()
    last_error = None

    for attempt in range(1, max_attempts + 1):
        try:
            print(f"[INFO] {supplier_name}: bandymas {attempt}/{max_attempts}")

            response = session.get(
                url,
                headers=headers,
                timeout=(20, 180),
                allow_redirects=True,
            )
            response.raise_for_status()

            content = response.content
            content_type = response.headers.get("Content-Type", "")

            print(f"[DEBUG] {supplier_name}: HTTP {response.status_code}, Content-Type={content_type}")

            if expect_xml and not looks_like_xml(content):
                debug_path = target_path.with_suffix(".debug.html")
                debug_path.write_bytes(content)
                raise ValueError(
                    f"{supplier_name} negrąžino XML. Debug failas: {debug_path}"
                )

            target_path.write_bytes(content)
            print(f"[OK] {supplier_name}: raw failas išsaugotas {target_path}")
            return content

        except Exception as e:
            last_error = e
            print(f"[WARN] {supplier_name}: nepavyko ({e})")
            if attempt < max_attempts:
                sleep_seconds = attempt * 15
                print(f"[INFO] {supplier_name}: laukiu {sleep_seconds}s...")
                time.sleep(sleep_seconds)

    raise RuntimeError(f"{supplier_name}: nepavyko parsiųsti po {max_attempts} bandymų: {last_error}")


def parse_marini(xml_bytes: bytes) -> list[dict[str, str]]:
    root = ET.fromstring(xml_bytes)
    rows: list[dict[str, str]] = []

    for item in root.findall(".//b2b"):
        kodas = clean_text(item.findtext("kod"))
        ean = clean_text(item.findtext("EAN"))
        likutis = clean_text(item.findtext("stan"))

        if not kodas and not ean and not likutis:
            continue

        rows.append({
            "Kodas": kodas,
            "EAN": ean,
            "Likutis": likutis,
        })

    return rows


def parse_zuja(xml_bytes: bytes) -> list[dict[str, str]]:
    root = ET.fromstring(xml_bytes)
    rows: list[dict[str, str]] = []

    for item in root.findall(".//product"):
        kodas = clean_text(item.findtext("sku"))
        ean = clean_text(item.findtext("barcode"))
        likutis = clean_text(item.findtext("total_quantity"))

        if not kodas and not ean and not likutis:
            continue

        rows.append({
            "Kodas": kodas,
            "EAN": ean,
            "Likutis": likutis,
        })

    return rows


def normalize_header(header: object) -> str:
    return clean_text(header).strip().lower()


def parse_tylla(xlsx_bytes: bytes) -> list[dict[str, str]]:
    if load_workbook is None:
        raise ImportError("Trūksta openpyxl. Paleisk: pip install openpyxl")

    workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []

    header_map: dict[str, int] = {}
    for idx, header in enumerate(header_row):
        normalized = normalize_header(header)
        if normalized:
            header_map[normalized] = idx

    required_columns = {
        "sku": "Kodas",
        "ean": "EAN",
        "stock_local": "Likutis",
    }

    missing = [col for col in required_columns if col not in header_map]
    if missing:
        raise ValueError(f"Tylla faile trūksta stulpelių: {', '.join(missing)}")

    rows: list[dict[str, str]] = []

    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        kodas = clean_text(excel_row[header_map["sku"]])
        ean = clean_text(excel_row[header_map["ean"]])
        likutis = clean_text(excel_row[header_map["stock_local"]])

        if not kodas and not ean and not likutis:
            continue

        rows.append({
            "Kodas": kodas,
            "EAN": ean,
            "Likutis": likutis,
        })

    return rows


def save_csv(supplier_name: str, rows: list[dict[str, str]], timestamp: str) -> Path:
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    csv_path = CSV_DIR / build_filename(supplier_name, ".csv", timestamp)

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["Kodas", "EAN", "Likutis"])
        writer.writeheader()
        writer.writerows(rows)

    return csv_path


def process_supplier(supplier_name: str, config: dict[str, str]) -> None:
    timestamp = current_timestamp()
    raw_path = OUTPUT_DIR / build_filename(supplier_name, config["raw_extension"], timestamp)

    expect_xml = config["type"] == "xml"
    content = download_with_retries(
        supplier_name=supplier_name,
        url=config["url"],
        target_path=raw_path,
        expect_xml=expect_xml,
    )

    if supplier_name == "Marini":
        rows = parse_marini(content)
    elif supplier_name == "Zuja":
        rows = parse_zuja(content)
    elif supplier_name == "Tylla":
        rows = parse_tylla(content)
    else:
        raise ValueError(f"Nežinomas tiekėjas: {supplier_name}")

    csv_path = save_csv(supplier_name, rows, timestamp)
    print(f"[OK] {supplier_name}: CSV išsaugotas {csv_path} ({len(rows)} eilučių)")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CSV_DIR.mkdir(parents=True, exist_ok=True)

    for supplier_name, config in SUPPLIERS.items():
        try:
            process_supplier(supplier_name, config)
        except Exception as e:
            print(f"[ERROR] {supplier_name}: {e}")


if __name__ == "__main__":
    main()
